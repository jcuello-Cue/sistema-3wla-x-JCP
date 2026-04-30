import streamlit as st
import json
import os
from datetime import datetime, date, timedelta
from openpyxl import load_workbook
import anthropic

# ─────────────────────────────────────────────
# CONFIGURACIÓN GENERAL
# ─────────────────────────────────────────────
STATE_FILE = "estado_3wla.json"
FIRMA = "Atte.\n\nJCP"

st.set_page_config(
    page_title="3WLA — LM5 Río Blanco",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# CSS personalizado
st.markdown("""
<style>
    .block-container { padding-top: 1.5rem; }
    .stTabs [data-baseweb="tab"] { font-size: 16px; font-weight: 600; padding: 12px 24px; }
    .stTabs [data-baseweb="tab-list"] { gap: 8px; }
    .metric-box {
        background: #f0f2f6; border-radius: 12px;
        padding: 16px; text-align: center; margin-bottom: 8px;
    }
    .metric-title { font-size: 13px; color: #666; margin-bottom: 4px; }
    .metric-value { font-size: 26px; font-weight: 700; color: #1f2937; }
    .metric-sub { font-size: 12px; color: #888; margin-top: 2px; }
    .act-card {
        background: #fff; border: 1px solid #e5e7eb;
        border-radius: 10px; padding: 14px 18px; margin-bottom: 10px;
    }
    .act-nombre { font-size: 15px; font-weight: 600; color: #1f2937; }
    .act-meta { font-size: 13px; color: #6b7280; margin-top: 3px; }
    .badge-area {
        display: inline-block; background: #e0e7ff; color: #3730a3;
        border-radius: 6px; padding: 2px 8px; font-size: 11px;
        font-weight: 600; margin-bottom: 6px;
    }
    .pendiente-tag {
        display: inline-block; background: #fef3c7; color: #92400e;
        border-radius: 6px; padding: 2px 8px; font-size: 11px;
        font-weight: 600; margin-left: 6px;
    }
    .avance-bar-bg {
        background: #e5e7eb; border-radius: 999px;
        height: 18px; width: 100%; overflow: hidden;
    }
    .email-box {
        background: #f9fafb; border: 1px solid #d1d5db;
        border-radius: 8px; padding: 16px; font-family: monospace;
        font-size: 13px; white-space: pre-wrap; color: #1f2937;
    }
    div[data-testid="stTextArea"] textarea {
        font-size: 13px; font-family: 'Courier New', monospace;
    }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# LECTURA DEL TRISEMANAL
# ─────────────────────────────────────────────
def leer_trisemanal(filepath):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["Programación"]
    rows = list(ws.iter_rows(min_row=1, max_row=7, values_only=True))

    r4 = rows[3]
    num_trisemanal = None
    fechas_encontradas = []
    for j, v in enumerate(r4):
        if v == "TRISEMANAL N°":
            num_trisemanal = r4[j + 1] if j + 1 < len(r4) else None
        if isinstance(v, datetime):
            fechas_encontradas.append(v.date())

    r7 = rows[6]
    fechas_s1 = []
    for j in range(32, 39):
        v = r7[j]
        if isinstance(v, datetime):
            fechas_s1.append(v.date())

    hh_totales_s1 = None
    hh_diarias_s1 = []  # totales por día según fila HH TOTALES
    actividades = []

    for row in ws.iter_rows(min_row=8, max_row=500, values_only=True):
        corr    = row[1]
        area    = row[11]
        nombre  = row[13]
        unidad  = row[15]
        rend    = row[16]
        hh_s1   = row[39]
        cant_s1 = row[40]
        dur     = row[22]
        inicio  = row[20]
        termino = row[21]

        if nombre and str(nombre).strip() == "HH TOTALES":
            hh_totales_s1 = hh_s1 if isinstance(hh_s1, (int, float)) else None
            # Guardar totales diarios exactos de las celdas naranjas
            hh_diarias_s1 = [
                row[j] if isinstance(row[j], (int, float)) else 0
                for j in range(32, 39)
            ]
            continue

        if not isinstance(corr, (int, float)):
            continue
        if not isinstance(hh_s1, (int, float)) or hh_s1 <= 0:
            continue

        cant_s1_v = cant_s1 if isinstance(cant_s1, (int, float)) else 0
        rend_v    = rend if isinstance(rend, (int, float)) else 0

        ini_date  = inicio.date() if isinstance(inicio, datetime) else None
        ter_date  = termino.date() if isinstance(termino, datetime) else None

        # Días activos en S1 = días de S1 que caen dentro del rango de la actividad
        dias_activos = []
        for idx, fd in enumerate(fechas_s1):
            if ini_date and ter_date:
                if ini_date <= fd <= ter_date:
                    dias_activos.append(idx)

        if not dias_activos:
            continue

        # CORRECCIÓN: dividir por días activos EN S1, no por duración total
        n_dias_s1 = len(dias_activos)
        cant_dia  = round(cant_s1_v / n_dias_s1, 6) if n_dias_s1 > 0 else 0
        hh_dia    = round(hh_s1 / n_dias_s1, 4) if n_dias_s1 > 0 else 0

        area_clean = str(area).strip() if area else "General"
        area_clean = area_clean.replace("Tramo ", "TRAMO ").upper() if area else "GENERAL"

        actividades.append({
            "corr":        int(corr),
            "area":        area_clean,
            "nombre":      str(nombre).strip(),
            "unidad":      str(unidad).strip() if unidad else "gl",
            "rendimiento": round(rend_v, 4),
            "cant_s1":     round(cant_s1_v, 6),
            "n_dias_s1":   n_dias_s1,
            "cant_dia":    cant_dia,
            "hh_dia":      hh_dia,
            "hh_s1":       round(hh_s1, 2),
            "inicio":      ini_date.isoformat() if ini_date else None,
            "termino":     ter_date.isoformat() if ter_date else None,
            "dias_activos": dias_activos,
            "fechas_s1":   [d.isoformat() for d in fechas_s1],
        })

    return {
        "num_trisemanal":  str(num_trisemanal) if num_trisemanal else "—",
        "fecha_inicio_s1": fechas_s1[0].isoformat() if fechas_s1 else None,
        "fecha_fin_s1":    fechas_s1[-1].isoformat() if fechas_s1 else None,
        "fechas_s1":       [d.isoformat() for d in fechas_s1],
        "hh_totales_s1":   round(hh_totales_s1, 2) if hh_totales_s1 else 0,
        "hh_diarias_s1":   [round(h, 2) for h in hh_diarias_s1],
        "actividades":     actividades,
    }

# ─────────────────────────────────────────────
# GESTIÓN DE ESTADO (GitHub como base de datos)
# ─────────────────────────────────────────────
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")
GITHUB_REPO  = os.environ.get("GITHUB_REPO", "jcuello-Cue/sistema-3wla-x-JCP")
GITHUB_FILE  = "estado_3wla.json"
GITHUB_API   = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_FILE}"

def cargar_estado():
    # Intentar desde GitHub si hay token
    if GITHUB_TOKEN:
        try:
            import urllib.request, base64
            req = urllib.request.Request(
                GITHUB_API,
                headers={
                    "Authorization": f"token {GITHUB_TOKEN}",
                    "Accept": "application/vnd.github.v3+json"
                }
            )
            with urllib.request.urlopen(req) as resp:
                data = json.loads(resp.read())
                content_b64 = data["content"].replace("\n","")
                estado = json.loads(base64.b64decode(content_b64).decode("utf-8"))
                # Guardar SHA para updates
                st.session_state["github_sha"] = data["sha"]
                return estado
        except Exception:
            pass
    # Fallback: archivo local
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return None

def guardar_estado(estado):
    # Guardar local siempre
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(estado, f, ensure_ascii=False, indent=2)
    # Subir a GitHub si hay token
    if GITHUB_TOKEN:
        try:
            import urllib.request, base64
            content_b64 = base64.b64encode(
                json.dumps(estado, ensure_ascii=False, indent=2).encode("utf-8")
            ).decode("utf-8")
            sha = st.session_state.get("github_sha", "")
            payload = {
                "message": "Actualización estado 3WLA",
                "content": content_b64,
                "branch": "main"
            }
            if sha:
                payload["sha"] = sha
            data = json.dumps(payload).encode("utf-8")
            req = urllib.request.Request(
                GITHUB_API,
                data=data,
                method="PUT",
                headers={
                    "Authorization": f"token {GITHUB_TOKEN}",
                    "Accept": "application/vnd.github.v3+json",
                    "Content-Type": "application/json"
                }
            )
            with urllib.request.urlopen(req) as resp:
                result = json.loads(resp.read())
                st.session_state["github_sha"] = result["content"]["sha"]
        except Exception as e:
            pass  # Si falla GitHub, igual quedó guardado local

# ─────────────────────────────────────────────
# LÓGICA DE CÁLCULO
# ─────────────────────────────────────────────
def actividades_del_dia(estado, fecha_str):
    """
    Retorna actividades activas hoy O que tienen pendiente acumulado de días anteriores.
    Aunque la actividad no esté programada hoy, si tiene deuda se muestra igual.
    """
    fd = date.fromisoformat(fecha_str)
    pendientes = pendientes_acumulados(estado, fecha_str)
    acts = []
    for a in estado["trisemanal"]["actividades"]:
        if not a["inicio"] or not a["termino"]:
            continue
        ini = date.fromisoformat(a["inicio"])
        ter = date.fromisoformat(a["termino"])
        # Mostrar si: hoy está en rango O tiene pendiente acumulado
        en_rango   = ini <= fd <= ter
        con_deuda  = pendientes.get(a["corr"], 0) > 0.0001
        if en_rango or con_deuda:
            acts.append(a)
    return acts

def pendientes_acumulados(estado, fecha_str):
    """
    Calcula pendiente acumulado neto por actividad.
    Considera TODOS los días registrados en S1 hasta ayer,
    incluyendo ejecuciones fuera del rango original de la actividad.
    """
    pendientes = {}
    fd = date.fromisoformat(fecha_str)
    registro = estado.get("registro", {})
    fechas_s1 = [date.fromisoformat(d) for d in estado["trisemanal"]["fechas_s1"]]

    for a in estado["trisemanal"]["actividades"]:
        if not a["inicio"] or not a["termino"]:
            continue
        ini = date.fromisoformat(a["inicio"])
        ter = date.fromisoformat(a["termino"])

        # PASO 1: Calcular HH esperadas (solo días en rango de la actividad)
        hh_esperadas = 0.0
        for fs in fechas_s1:
            if fs < fd and ini <= fs <= ter:
                hh_esperadas += a["hh_dia"]

        # PASO 2: Calcular HH ejecutadas de esta actividad
        # en TODOS los días anteriores (aunque estén fuera del rango)
        hh_ejecutadas = 0.0
        for fs in fechas_s1:
            if fs < fd:
                act_reg = registro.get(fs.isoformat(), {}).get(str(a["corr"]))
                if act_reg is not None:
                    hh_ejecutadas += act_reg.get("cant_ejecutada", 0) * a["rendimiento"]

        # PASO 3: Pendiente neto en HH → convertir a unidades
        hh_pendiente = round(hh_esperadas - hh_ejecutadas, 4)

        if hh_pendiente > 0.01:
            # Convertir HH pendientes a unidades para mostrar en pantalla
            cant_pendiente = round(hh_pendiente / a["rendimiento"], 6) if a["rendimiento"] > 0 else 0
            pendientes[a["corr"]] = cant_pendiente
        elif hh_pendiente < -0.01:
            # Excedente neto
            cant_excedente = round(hh_pendiente / a["rendimiento"], 6) if a["rendimiento"] > 0 else 0
            pendientes[a["corr"]] = cant_excedente

    return pendientes

def calcular_acumulado(estado, hasta_fecha_str=None):
    """
    Calcula acumulado S1:
    - hh_esperadas: suma de HH planificadas desde inicio S1 hasta hoy inclusive
    - hh_ejecutadas: suma de HH realmente ejecutadas desde inicio S1 hasta hoy inclusive
      (usa cant_ejecutada * rendimiento de cada registro diario)
    """
    hh_ej_total  = 0.0
    hh_esp_total = 0.0
    por_dia  = {}
    por_area = {}

    tri       = estado["trisemanal"]
    registro  = estado.get("registro", {})
    hoy       = date.fromisoformat(hasta_fecha_str) if hasta_fecha_str else date.today()
    fechas_s1 = tri["fechas_s1"]

    for fecha_str in fechas_s1:
        fd = date.fromisoformat(fecha_str)
        if fd > hoy:
            por_dia[fecha_str] = {"esperado": 0, "ejecutado": 0, "registrado": False}
            continue

        hh_ej_dia  = 0.0
        hh_esp_dia = 0.0

        # HH ESPERADAS: suma de actividades individuales (misma fuente que déficits)
        for a in tri["actividades"]:
            if not a["inicio"] or not a["termino"]:
                continue
            ini = date.fromisoformat(a["inicio"])
            ter = date.fromisoformat(a["termino"])
            if not (ini <= fd <= ter):
                continue
            hh_esp_dia += a["hh_dia"]
            area = a["area"]
            if area not in por_area:
                por_area[area] = {"esperado": 0, "ejecutado": 0}
            por_area[area]["esperado"] += a["hh_dia"]

        # HH EJECUTADAS: solo de fechas dentro del trisemanal actual
        if fecha_str in fechas_s1:
            reg_dia = registro.get(fecha_str, {})
            for corr_str, act_reg in reg_dia.items():
                cant_ej = act_reg.get("cant_ejecutada", 0)
                rend    = act_reg.get("rendimiento", 0)
                hh_ej   = round(cant_ej * rend, 2)
                hh_ej_dia += hh_ej
                area = act_reg.get("area", "Sin área")
                if area not in por_area:
                    por_area[area] = {"esperado": 0, "ejecutado": 0}
                por_area[area]["ejecutado"] += hh_ej

        por_dia[fecha_str] = {
            "esperado":   round(hh_esp_dia, 2),
            "ejecutado":  round(hh_ej_dia, 2),
            "registrado": fecha_str in registro
        }
        hh_ej_total  += hh_ej_dia
        hh_esp_total += hh_esp_dia

    # Calcular por_resp: déficit neto por responsable
    por_resp = {}
    fechas_dates = [date.fromisoformat(d) for d in fechas_s1]
    for a in tri["actividades"]:
        if not a["inicio"] or not a["termino"]: continue
        ini_r = date.fromisoformat(a["inicio"])
        ter_r = date.fromisoformat(a["termino"])
        hh_esp_r = sum(a["hh_dia"] for fd in fechas_dates if fd <= hoy and ini_r <= fd <= ter_r)
        hh_ej_r  = sum(
            (registro.get(fd.isoformat(),{}).get(str(a["corr"]),{}).get("cant_ejecutada",0) or 0) * a["rendimiento"]
            for fd in fechas_dates if fd <= hoy
        )
        deficit_r = round(hh_esp_r - hh_ej_r, 2)
        if deficit_r > 0.5:
            resp_r = "Sin asignar"
            for fd in sorted(fechas_dates, reverse=True):
                if fd > hoy: continue
                reg_r = registro.get(fd.isoformat(),{}).get(str(a["corr"]))
                if reg_r and reg_r.get("responsable"):
                    resp_r = reg_r["responsable"]; break
            por_resp[resp_r] = por_resp.get(resp_r, 0) + deficit_r

    # Si hh_totales_s1 es 0, calcularlo como suma de todas las HH de S1
    hh_meta = tri["hh_totales_s1"]
    if not hh_meta or hh_meta == 0:
        hh_meta = sum(a["hh_s1"] for a in tri["actividades"] if isinstance(a.get("hh_s1"), (int,float)))
        hh_meta = round(hh_meta, 2)

    return {
        "hh_ejecutadas": round(hh_ej_total, 2),
        "hh_esperadas":  round(hh_esp_total, 2),
        "hh_meta":       hh_meta,
        "por_dia":       por_dia,
        "por_area":      por_area,
        "por_resp":      por_resp,
    }

# ─────────────────────────────────────────────
# PANEL DE ACUMULADO (reutilizable)
# ─────────────────────────────────────────────
def panel_acumulado(estado, fecha_str, tab_key="a"):
    import pandas as pd

    acu     = calcular_acumulado(estado, fecha_str)
    hh_ej   = acu["hh_ejecutadas"]
    hh_esp  = acu["hh_esperadas"]
    hh_meta = acu["hh_meta"]
    pct_ej  = round(hh_ej / hh_meta * 100, 1) if hh_meta else 0
    pct_esp = round(hh_esp / hh_meta * 100, 1) if hh_meta else 0

    st.markdown("---")
    st.markdown("### 📊 Acumulado S1")

    # Métricas
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("✅ HH Ejecutadas", f"{hh_ej:,.1f}", f"{pct_ej}% de la meta")
    with c2:
        st.metric("📅 HH Esperadas a hoy", f"{hh_esp:,.1f}", f"{pct_esp}% de la meta")
    with c3:
        delta = round(hh_ej - hh_esp, 1)
        st.metric("📈 Diferencia", f"{delta:+,.1f} HH",
                  "Adelantado" if delta >= 0 else "Atrasado",
                  delta_color="normal" if delta >= 0 else "inverse")
    with c4:
        st.metric("🎯 Meta S1", f"{hh_meta:,.1f} HH",
                  f"Faltan {round(hh_meta - hh_ej, 1):,.1f} HH")

    st.progress(pct_ej / 100, text=f"Avance: {pct_ej}%")

    # Botón PowerPoint
    col_ppt1, col_ppt2 = st.columns([2,3])
    with col_ppt1:
        if st.button("📊 Exportar resumen PowerPoint", type="primary",
                     use_container_width=True, key=f"btn_ppt_{tab_key}"):
            output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "resumen_3wla.pptx")
            script_path = os.path.join(os.getcwd(), "generar_ppt_python.py")
            with st.spinner("Generando PowerPoint..."):
                try:
                    import importlib.util, sys as _sys
                    spec = importlib.util.spec_from_file_location("generar_ppt_python", script_path)
                    mod  = importlib.util.module_from_spec(spec)
                    spec.loader.exec_module(mod)
                    mod.generar_ppt(STATE_FILE, fecha_str, output_path)
                    with open(output_path, "rb") as f:
                        st.session_state[f"ppt_bytes_{tab_key}"] = f.read()
                    st.session_state[f"ppt_ready_{tab_key}"] = True
                    st.rerun()
                except ModuleNotFoundError:
                    st.error("Instala python-pptx: python -m pip install python-pptx")
                except Exception as e:
                    st.error(f"Error generando PPT: {e}")

    if st.session_state.get(f"ppt_ready_{tab_key}"):
        with col_ppt2:
            st.download_button(
                label="⬇️ Descargar PowerPoint",
                data=st.session_state[f"ppt_bytes_{tab_key}"],
                file_name=f"Resumen_3WLA_S1_{fecha_str}.pptx",
                mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                key=f"dl_ppt_{tab_key}"
            )

    try:
        import plotly.graph_objects as go
        import plotly.express as px

        tri       = estado["trisemanal"]
        fechas_s1 = tri["fechas_s1"]
        registro  = estado.get("registro", {})
        hoy       = date.fromisoformat(fecha_str)
        fechas_s1_dates = [date.fromisoformat(d) for d in fechas_s1]
        dias_labels = [date.fromisoformat(d).strftime("%a %d/%m") for d in fechas_s1]

        esp_dias = [acu["por_dia"].get(d, {}).get("esperado", 0)  for d in fechas_s1]
        ej_dias  = [acu["por_dia"].get(d, {}).get("ejecutado", 0) for d in fechas_s1]

        # ── GRÁFICO 1: HH por día ────────────────────────────────
        fig1 = go.Figure()
        fig1.add_trace(go.Bar(
            name="Esperado", x=dias_labels, y=esp_dias,
            marker_color="#93c5fd",
            text=[f"{v:,.0f}" for v in esp_dias],
            textposition="outside", textfont=dict(size=12)
        ))
        fig1.add_trace(go.Bar(
            name="Ejecutado", x=dias_labels, y=ej_dias,
            marker_color="#1d4ed8",
            text=[f"{v:,.0f}" for v in ej_dias],
            textposition="outside", textfont=dict(size=12)
        ))
        fig1.update_layout(
            title="① HH por día — Esperado vs Ejecutado",
            barmode="group", height=420,
            legend=dict(orientation="h", y=-0.2),
            yaxis_title="HH", xaxis_title="",
            margin=dict(t=50, b=60)
        )

        # ── GRÁFICO 2: Tendencia acumulada ───────────────────────
        acu_esp = []
        acu_ej  = []
        s_esp = s_ej = 0
        for e, j in zip(esp_dias, ej_dias):
            s_esp += e; acu_esp.append(round(s_esp,1))
            s_ej  += j; acu_ej.append(round(s_ej,1))

        fig2 = go.Figure()
        fig2.add_trace(go.Scatter(
            name="Acumulado Esperado", x=dias_labels, y=acu_esp,
            mode="lines+markers+text",
            line=dict(color="#93c5fd", width=3),
            marker=dict(size=8),
            text=[f"{v:,.0f}" for v in acu_esp],
            textposition="top center", textfont=dict(size=11)
        ))
        fig2.add_trace(go.Scatter(
            name="Acumulado Ejecutado", x=dias_labels, y=acu_ej,
            mode="lines+markers+text",
            line=dict(color="#1d4ed8", width=3),
            marker=dict(size=8),
            text=[f"{v:,.0f}" for v in acu_ej],
            textposition="bottom center", textfont=dict(size=11)
        ))
        fig2.update_layout(
            title="② Tendencia acumulada S1",
            height=420, legend=dict(orientation="h", y=-0.2),
            yaxis_title="HH acumuladas", xaxis_title="",
            margin=dict(t=50, b=60)
        )

        col_g1, col_g2 = st.columns(2)
        with col_g1:
            st.plotly_chart(fig1, use_container_width=True, key=f"chart_1_{tab_key}")
        with col_g2:
            st.plotly_chart(fig2, use_container_width=True, key=f"chart_2_{tab_key}")

        # ── GRÁFICO 3: Por actividad ─────────────────────────────
        rows_act = []
        for a in tri["actividades"]:
            ini = date.fromisoformat(a["inicio"]) if a["inicio"] else None
            ter = date.fromisoformat(a["termino"]) if a["termino"] else None
            if not ini or not ter:
                continue
            hh_esp_act = 0.0
            hh_ej_act  = 0.0
            cursor = ini
            while cursor <= min(ter, hoy):
                if cursor in fechas_s1_dates:
                    hh_esp_act += a["hh_dia"]
                    reg = registro.get(cursor.isoformat(), {}).get(str(a["corr"]))
                    if reg:
                        hh_ej_act += round(reg.get("cant_ejecutada", 0) * a["rendimiento"], 2)
                cursor += timedelta(days=1)
            if hh_esp_act > 0 or hh_ej_act > 0:
                nombre_corto = a["nombre"][:40] + "…" if len(a["nombre"]) > 40 else a["nombre"]
                rows_act.append({
                    "Actividad": nombre_corto,
                    "Esperado":  round(hh_esp_act, 1),
                    "Ejecutado": round(hh_ej_act, 1),
                    "Cumplimiento": f"{round(hh_ej_act/hh_esp_act*100,1) if hh_esp_act else 0}%"
                })

        if rows_act:
            df_act = pd.DataFrame(rows_act).sort_values("Esperado", ascending=True)
            fig3 = go.Figure()
            fig3.add_trace(go.Bar(
                name="Esperado", y=df_act["Actividad"], x=df_act["Esperado"],
                orientation="h", marker_color="#93c5fd",
                text=[f"{v:,.0f} HH" for v in df_act["Esperado"]],
                textposition="auto", textfont=dict(size=11)
            ))
            fig3.add_trace(go.Bar(
                name="Ejecutado", y=df_act["Actividad"], x=df_act["Ejecutado"],
                orientation="h", marker_color="#1d4ed8",
                text=[f"{v:,.0f} HH" for v in df_act["Ejecutado"]],
                textposition="auto", textfont=dict(size=11)
            ))
            fig3.update_layout(
                title="③ HH por actividad — Esperado vs Ejecutado",
                barmode="group", height=max(400, len(rows_act)*45),
                legend=dict(orientation="h", y=-0.1),
                xaxis_title="HH", yaxis_title="",
                margin=dict(t=50, b=60, l=20)
            )
            st.plotly_chart(fig3, use_container_width=True, key=f"chart_3_{tab_key}")

            # Tabla actividades
            df_act_tabla = df_act[["Actividad","Esperado","Ejecutado","Cumplimiento"]].sort_values("Esperado", ascending=False)
            st.dataframe(df_act_tabla, use_container_width=True, hide_index=True)

        # ── GRÁFICO 4: Por tramo/área ─────────────────────────────
        if acu["por_area"]:
            rows_area = []
            for area, v in sorted(acu["por_area"].items()):
                pct = round(v["ejecutado"] / v["esperado"] * 100, 1) if v["esperado"] > 0 else 0
                rows_area.append({
                    "Tramo/Área":   area,
                    "Esperado":     round(v["esperado"], 1),
                    "Ejecutado":    round(v["ejecutado"], 1),
                    "Cumplimiento": f"{pct}%"
                })
            df_area = pd.DataFrame(rows_area)
            fig4 = go.Figure()
            fig4.add_trace(go.Bar(
                name="Esperado", x=df_area["Tramo/Área"], y=df_area["Esperado"],
                marker_color="#93c5fd",
                text=[f"{v:,.0f}" for v in df_area["Esperado"]],
                textposition="outside", textfont=dict(size=13)
            ))
            fig4.add_trace(go.Bar(
                name="Ejecutado", x=df_area["Tramo/Área"], y=df_area["Ejecutado"],
                marker_color="#1d4ed8",
                text=[f"{v:,.0f}" for v in df_area["Ejecutado"]],
                textposition="outside", textfont=dict(size=13)
            ))
            fig4.update_layout(
                title="④ HH por Tramo/Área — Esperado vs Ejecutado",
                barmode="group", height=420,
                legend=dict(orientation="h", y=-0.2),
                yaxis_title="HH", xaxis_title="",
                margin=dict(t=50, b=60)
            )
            col_g3, col_g4 = st.columns([3, 1])
            with col_g3:
                st.plotly_chart(fig4, use_container_width=True, key=f"chart_4_{tab_key}")
            with col_g4:
                st.dataframe(
                    df_area[["Tramo/Área", "Esperado", "Ejecutado", "Cumplimiento"]],
                    use_container_width=True,
                    hide_index=True
                )

        # ── GRÁFICO 5: HH no ejecutadas por responsable ──────────
        st.markdown("**⑤ HH no ejecutadas por responsable (acumulado hasta hoy)**")

        # Calcular déficit NETO por actividad (misma lógica que email de cierre)
        # y atribuirlo al responsable del último registro con no cumplimiento
        resp_hh = {}
        for a_base in tri["actividades"]:
            if not a_base["inicio"] or not a_base["termino"]:
                continue
            ini_a = date.fromisoformat(a_base["inicio"])
            ter_a = date.fromisoformat(a_base["termino"])

            # HH esperadas hasta hoy
            hh_esp_act = sum(
                a_base["hh_dia"] for fs in fechas_s1_dates
                if fs <= hoy and ini_a <= fs <= ter_a
            )
            # HH ejecutadas en todos los días registrados
            hh_ej_act = sum(
                registro.get(fs.isoformat(), {}).get(str(a_base["corr"]), {}).get("cant_ejecutada", 0)
                * a_base["rendimiento"]
                for fs in fechas_s1_dates if fs <= hoy
            )
            deficit_neto = round(hh_esp_act - hh_ej_act, 2)

            if deficit_neto > 0.5:
                # Buscar responsable del último registro con no cumplimiento
                resp = "Sin asignar"
                for fs in sorted(fechas_s1_dates, reverse=True):
                    if fs <= hoy:
                        act_reg_r = registro.get(fs.isoformat(), {}).get(str(a_base["corr"]))
                        if act_reg_r and act_reg_r.get("responsable"):
                            resp = act_reg_r["responsable"]
                            break
                resp_hh[resp] = resp_hh.get(resp, 0) + deficit_neto

        if resp_hh:
            resp_sorted = sorted(resp_hh.items(), key=lambda x: -x[1])
            labels_resp = [r[0] for r in resp_sorted]
            values_resp = [round(r[1], 1) for r in resp_sorted]
            total_resp  = sum(values_resp)

            colors_map = {
                "Fe Grande":      "#ef4444",
                "R&Q Ingeniería": "#f97316",
                "CODELCO":        "#eab308",
                "Externo":        "#8b5cf6",
                "Otro":           "#6b7280",
                "Sin asignar":    "#d1d5db",
            }

            col_r1, col_r2 = st.columns([2, 1])
            with col_r1:
                fig5 = go.Figure()
                fig5.add_trace(go.Bar(
                    x=labels_resp,
                    y=values_resp,
                    marker_color=[colors_map.get(r, "#6b7280") for r in labels_resp],
                    text=[f"{v:,.1f} HH  ({round(v/total_resp*100,1)}%)" for v in values_resp],
                    textposition="outside",
                    textfont=dict(size=13),
                ))
                fig5.update_layout(
                    title="⑤ HH no ejecutadas por responsable",
                    height=420,
                    yaxis_title="HH no ejecutadas",
                    xaxis_title="",
                    showlegend=False,
                    margin=dict(t=50, b=80)
                )
                st.plotly_chart(fig5, use_container_width=True, key=f"chart_5_{tab_key}")

            with col_r2:
                rows_resp = [{"Responsable": r, "HH no ejecutadas": round(h,1), "% del total": f"{round(h/total_resp*100,1)}%"} for r,h in resp_sorted]
                rows_resp.append({"Responsable": "TOTAL", "HH no ejecutadas": round(total_resp,1), "% del total": "100%"})
                st.dataframe(pd.DataFrame(rows_resp), use_container_width=True, hide_index=True)
            # ── TABLA DETALLADA DE JUSTIFICACIONES ───────────────
            st.markdown("---")
            st.markdown("**📋 Detalle de HH no ejecutadas con justificaciones**")

            dias_labels_map = {fs.isoformat(): fs.strftime("%a %d/%m") for fs in fechas_s1_dates}
            rows_detalle = []

            for a_base in tri["actividades"]:
                if not a_base["inicio"] or not a_base["termino"]:
                    continue
                ini_a = date.fromisoformat(a_base["inicio"])
                ter_a = date.fromisoformat(a_base["termino"])

                for fs in sorted(fechas_s1_dates):
                    if fs > hoy:
                        continue
                    act_reg_r = registro.get(fs.isoformat(), {}).get(str(a_base["corr"]))
                    if not act_reg_r:
                        continue
                    cant_ej = act_reg_r.get("cant_ejecutada", 0)
                    hh_ej   = round(cant_ej * a_base["rendimiento"], 2)
                    hh_esp  = a_base["hh_dia"] if ini_a <= fs <= ter_a else 0
                    deficit = round(hh_esp - hh_ej, 2)
                    causa   = act_reg_r.get("causa", "")
                    resp    = act_reg_r.get("responsable", "Sin asignar")
                    if deficit > 0.5 or (causa and hh_ej < hh_esp):
                        rows_detalle.append({
                            "Fecha":         dias_labels_map.get(fs.isoformat(), fs.isoformat()),
                            "Área":          a_base["area"],
                            "Actividad":     a_base["nombre"],
                            "HH Esperadas":  round(hh_esp, 1),
                            "HH Ejecutadas": round(hh_ej, 1),
                            "HH Déficit":    round(deficit, 1),
                            "Responsable":   resp,
                            "Justificación": causa if causa else "Sin justificación registrada",
                        })

            if rows_detalle:
                df_det = pd.DataFrame(rows_detalle)
                responsables_unicos = ["Todos"] + sorted(df_det["Responsable"].unique().tolist())
                filtro_resp = st.selectbox("Filtrar por responsable:", responsables_unicos, key=f"filtro_responsable_{tab_key}")
                if filtro_resp != "Todos":
                    df_det = df_det[df_det["Responsable"] == filtro_resp]
                # Mostrar como HTML para que el texto se vea completo
                import html as _html
                html_rows = ""
                for _, row in df_det.iterrows():
                    html_rows += f"""
                    <tr style="border-bottom:1px solid #E5E7EB;vertical-align:top">
                        <td style="padding:8px;font-size:12px;white-space:nowrap">{row['Fecha']}</td>
                        <td style="padding:8px;font-size:12px;white-space:nowrap">{row['Área']}</td>
                        <td style="padding:8px;font-size:12px">{_html.escape(str(row['Actividad']))}</td>
                        <td style="padding:8px;font-size:12px;text-align:center">{row['HH Esperadas']:.1f}</td>
                        <td style="padding:8px;font-size:12px;text-align:center">{row['HH Ejecutadas']:.1f}</td>
                        <td style="padding:8px;font-size:12px;text-align:center;color:#EF4444;font-weight:600">{row['HH Déficit']:.1f}</td>
                        <td style="padding:8px;font-size:12px;white-space:nowrap">{_html.escape(str(row['Responsable']))}</td>
                        <td style="padding:8px;font-size:12px;line-height:1.5">{_html.escape(str(row['Justificación']))}</td>
                    </tr>"""

                st.markdown(f"""
                <div style="overflow-x:auto">
                <table style="width:100%;border-collapse:collapse;font-family:sans-serif">
                    <thead>
                        <tr style="background:#1E3A5F;color:white">
                            <th style="padding:8px;font-size:11px;text-align:left">Fecha</th>
                            <th style="padding:8px;font-size:11px;text-align:left">Área</th>
                            <th style="padding:8px;font-size:11px;text-align:left">Actividad</th>
                            <th style="padding:8px;font-size:11px;text-align:center">HH Esp.</th>
                            <th style="padding:8px;font-size:11px;text-align:center">HH Ej.</th>
                            <th style="padding:8px;font-size:11px;text-align:center">HH Déf.</th>
                            <th style="padding:8px;font-size:11px;text-align:left">Responsable</th>
                            <th style="padding:8px;font-size:11px;text-align:left">Justificación</th>
                        </tr>
                    </thead>
                    <tbody>{html_rows}</tbody>
                </table>
                </div>
                """, unsafe_allow_html=True)
                total_def = df_det["HH Déficit"].sum()
                st.caption(f"Total HH no ejecutadas mostradas: **{total_def:,.1f} HH** | Registros: {len(df_det)}")
            else:
                st.info("Sin registros de no cumplimiento en el período.")

        else:
            st.info("Sin HH no ejecutadas registradas hasta la fecha.")

        # ── TABLA RESUMEN FINAL S1 ───────────────────────────────
        st.markdown("---")
        st.markdown("### 📋 Actividades no cumplidas en S1 — Detalle por día")
        st.caption("Todas las actividades con déficit al cierre del período, con justificaciones de cada jornada.")

        # Construir tabla detallada por actividad y día
        fechas_s1_dates_local = [date.fromisoformat(d) for d in tri["fechas_s1"]]
        rows_final = []

        for a_f in tri["actividades"]:
            if not a_f["inicio"] or not a_f["termino"]: continue
            ini_f = date.fromisoformat(a_f["inicio"])
            ter_f = date.fromisoformat(a_f["termino"])

            # Calcular déficit neto total de la actividad
            hh_esp_total_f = 0.0
            hh_ej_total_f  = 0.0
            for fd_f in fechas_s1_dates_local:
                if ini_f <= fd_f <= ter_f:
                    hh_esp_total_f += a_f["hh_dia"]
                reg_f = registro.get(fd_f.isoformat(), {}).get(str(a_f["corr"]))
                if reg_f:
                    hh_ej_total_f += (reg_f.get("cant_ejecutada", 0) or 0) * a_f["rendimiento"]

            deficit_f = round(hh_esp_total_f - hh_ej_total_f, 1)
            if deficit_f <= 0.5:
                continue  # Actividad cumplida, skip

            # Recopilar justificaciones por día
            justificaciones = []
            for fd_f in sorted(fechas_s1_dates_local):
                reg_f = registro.get(fd_f.isoformat(), {}).get(str(a_f["corr"]))
                if not reg_f: continue
                cant_ej_f  = reg_f.get("cant_ejecutada", 0) or 0
                hh_ej_f    = round(cant_ej_f * a_f["rendimiento"], 1)
                hh_esp_f   = a_f["hh_dia"] if ini_f <= fd_f <= ter_f else 0
                deficit_dia = round(hh_esp_f - hh_ej_f, 1)
                causa_f    = reg_f.get("causa", "")
                resp_f     = reg_f.get("responsable", "")
                # Solo mostrar días con déficit real (no excedentes) o con causa registrada
                if deficit_dia > 0.5 or (causa_f and deficit_dia >= 0):
                    justificaciones.append(
                        f"{fd_f.strftime('%d/%m')}: "
                        f"{deficit_dia:.1f} HH no ejecutadas. "
                        f"{causa_f if causa_f else 'Sin justificación.'} "
                        f"[{resp_f}]"
                    )

            pct_cumpl = round(hh_ej_total_f / hh_esp_total_f * 100) if hh_esp_total_f > 0 else 0
            rows_final.append({
                "Área":         a_f["area"],
                "Actividad":    a_f["nombre"],
                "HH Déficit":   deficit_f,
                "% Cumpl.":     f"{pct_cumpl}%",
                "Justificaciones por día": " | ".join(justificaciones) if justificaciones else "Sin registros.",
            })

        if rows_final:
            import pandas as pd
            total_deficit_s1 = sum(r["HH Déficit"] for r in rows_final)
            st.metric("Total HH no ejecutadas S1",
                      f"{total_deficit_s1:,.1f} HH",
                      f"{round(total_deficit_s1/tri['hh_totales_s1']*100,1)}% de la meta",
                      delta_color="inverse")

            # Agrupar por área y mostrar tabla por cada una
            areas_orden = sorted(set(r["Área"] for r in rows_final))
            for area_f in areas_orden:
                acts_area_f = sorted(
                    [r for r in rows_final if r["Área"] == area_f],
                    key=lambda x: -x["HH Déficit"]
                )
                deficit_area = sum(r["HH Déficit"] for r in acts_area_f)

                st.markdown(
                    f"<div style='background:#1E3A5F;color:white;padding:8px 14px;"
                    f"border-radius:6px;font-weight:700;font-size:13px;margin-top:16px;"
                    f"display:flex;justify-content:space-between'>"
                    f"<span>{area_f}</span>"
                    f"<span style='color:#FCA5A5'>{deficit_area:,.1f} HH no ejecutadas</span>"
                    f"</div>",
                    unsafe_allow_html=True
                )

                for act_f in acts_area_f:
                    pct_color = "#DC2626" if int(act_f["% Cumpl."].replace("%","")) < 30 else (
                        "#D97706" if int(act_f["% Cumpl."].replace("%","")) < 70 else "#16A34A"
                    )
                    st.markdown(
                        f"<div style='border:1px solid #E5E7EB;border-radius:6px;"
                        f"padding:10px 14px;margin-top:6px;background:#FAFAFA'>"
                        f"<div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:6px'>"
                        f"<span style='font-weight:600;font-size:13px'>{act_f['Actividad']}</span>"
                        f"<span style='display:flex;gap:12px;font-size:12px'>"
                        f"<span style='color:#EF4444;font-weight:600'>Déficit: {act_f['HH Déficit']:,.1f} HH</span>"
                        f"<span style='color:{pct_color};font-weight:600'>Cumpl.: {act_f['% Cumpl.']}</span>"
                        f"</span></div>"
                        f"<div style='font-size:11.5px;color:#374151;line-height:1.7'>"
                        + "".join(
                            f"<div style='border-left:3px solid #E5E7EB;padding-left:8px;margin-bottom:4px'>"
                            f"<b style='color:#6B7280'>{j.split(':')[0]}:</b> "
                            f"{':'.join(j.split(':')[1:]).strip()}</div>"
                            for j in act_f["Justificaciones por día"].split(" | ") if j
                        )
                        + "</div></div>",
                        unsafe_allow_html=True
                    )
        else:
            st.success("✅ Todas las actividades de S1 fueron ejecutadas según lo planificado.")

    except Exception as e:
        st.warning(f"Error generando gráficos: {e}")

# ─────────────────────────────────────────────
# GENERACIÓN DE EMAILS
# ─────────────────────────────────────────────
def generar_email_inicio(estado, fecha_str, acts_hoy, pendientes):
    fd_hoy = date.fromisoformat(fecha_str)
    lineas_por_area = {}
    for a in acts_hoy:
        pend = pendientes.get(a["corr"], 0)
        pend_positivo = max(pend, 0.0)

        ini_act = date.fromisoformat(a["inicio"]) if a["inicio"] else fd_hoy
        ter_act = date.fromisoformat(a["termino"]) if a["termino"] else fd_hoy
        en_rango = ini_act <= fd_hoy <= ter_act

        if en_rango:
            cant_hoy  = round(a["cant_dia"], 4)
            cant_pend = round(pend_positivo, 4)
        else:
            cant_hoy  = 0.0
            cant_pend = round(pend_positivo, 4)

        cant_total = round(cant_hoy + cant_pend, 4)
        hh_total   = round(cant_total * a["rendimiento"], 2)

        if cant_total <= 0 or hh_total <= 0:
            continue

        area = a["area"]
        if area not in lineas_por_area:
            lineas_por_area[area] = []
        lineas_por_area[area].append({
            "nombre":    a["nombre"],
            "cant":      cant_total,
            "unidad":    a["unidad"],
            "hh":        hh_total,
            "cant_dia":  cant_hoy,
            "cant_pend": cant_pend,
            "es_del_dia": en_rango and cant_pend == 0,
            "tiene_pendiente": cant_pend > 0,
        })

    fd_fmt = datetime.strptime(fecha_str, "%Y-%m-%d").strftime("%d de %B de %Y")
    datos  = json.dumps(lineas_por_area, ensure_ascii=False, indent=2)

    prompt = f"""Genera un email formal en español para el equipo de construcción del proyecto LM5 Río Blanco, informando las HH comprometidas para la jornada de hoy ({fd_fmt}) según el programa 3WLA.

Actividades por área (con desglose):
{datos}

FORMATO EXACTO:
- Saludo: "Estimados,"
- Párrafo: "Junto con saludar, comparto para su consideración las HH comprometidas para la jornada de hoy, de acuerdo con lo establecido en el 3WLA, las cuales incluyen tanto las programadas para esta fecha como aquellas pendientes de ejecución provenientes de los días anteriores."
- Por cada área en MAYÚSCULAS y negrita
- Por actividad: • **Nombre:** X.XX unidad → HH HH
  - Si tiene_pendiente=true y cant_dia>0: agrega entre paréntesis "(X.XX unidad del día + X.XX unidad pendiente de días anteriores)"
  - Si es_del_dia=true: agrega entre paréntesis "(programado del día)"
  - Si cant_dia=0 y tiene_pendiente=true: agrega entre paréntesis "(pendiente acumulado de días anteriores)"
- Cierre: "Saludos cordiales."
- Firma: {FIRMA}

No incluyas actividades con 0 HH. Sé preciso y formal."""

    client = anthropic.Anthropic()
    resp = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1500,
        temperature=0,
        messages=[{"role": "user", "content": prompt}]
    )
    return resp.content[0].text

def generar_email_cierre(estado, fecha_str, registro_dia):
    tri      = estado["trisemanal"]
    registro = estado.get("registro", {})
    fd       = date.fromisoformat(fecha_str)
    fd_fmt   = datetime.strptime(fecha_str, "%Y-%m-%d").strftime("%d de %B de %Y")

    # Construir lista de TODAS las actividades del día con su estado
    actividades_email = {}  # area -> lista

    # Calcular déficit neto por actividad — misma lógica que calcular_acumulado
    fechas_s1_dates = [date.fromisoformat(d) for d in tri["fechas_s1"]]
    deficit_por_corr = {}
    for a in tri["actividades"]:
        if not a["inicio"] or not a["termino"]:
            continue
        ini = date.fromisoformat(a["inicio"])
        ter = date.fromisoformat(a["termino"])

        hh_esp_acum = 0.0
        hh_ej_acum  = 0.0

        # Esperado: suma de hh_dia por cada día en S1 hasta hoy dentro del rango
        for fs in fechas_s1_dates:
            if fs <= fd and ini <= fs <= ter:
                hh_esp_acum += a["hh_dia"]

        # Ejecutado: suma de cant_ejecutada × rendimiento de todos los registros
        for fs in fechas_s1_dates:
            if fs <= fd:
                reg_c = registro.get(fs.isoformat(), {}).get(str(a["corr"]))
                if reg_c is not None:
                    hh_ej_acum += reg_c.get("cant_ejecutada", 0) * a["rendimiento"]

        deficit_neto = round(hh_esp_acum - hh_ej_acum, 2)
        deficit_por_corr[str(a["corr"])] = max(deficit_neto, 0)

    for corr_str, reg in registro_dia.items():
        area      = reg.get("area", "Sin área")
        nombre    = reg.get("nombre", "")
        unidad    = reg.get("unidad", "")
        cat       = reg.get("categoria", "Planificada del día")
        cant_esp  = reg.get("cant_esperada", 0)
        cant_ej   = reg.get("cant_ejecutada", 0)
        hh_ej     = reg.get("hh_ejecutadas", 0)
        causa     = reg.get("causa", "")
        resp      = reg.get("responsable", "Fe Grande")

        # Usar déficit neto acumulado real (no HH comprometidas del día)
        hh_deficit = deficit_por_corr.get(corr_str, 0)
        ejecutado  = hh_deficit < 1.0  # menos de 1 HH de déficit = ejecutado

        if area not in actividades_email:
            actividades_email[area] = []
        actividades_email[area].append({
            "nombre":      nombre,
            "categoria":   cat,
            "unidad":      unidad,
            "cant_esp":    cant_esp,
            "cant_ej":     cant_ej,
            "hh_ej":       hh_ej,
            "hh_deficit":  hh_deficit,  # déficit neto real para el email
            "ejecutado":   ejecutado,
            "causa":       causa,
            "responsable": resp,
        })

    acu     = calcular_acumulado(estado, fecha_str)
    hh_ej_t = acu["hh_ejecutadas"]
    hh_esp_t = acu["hh_esperadas"]
    hh_meta  = acu["hh_meta"]
    datos    = json.dumps(actividades_email, ensure_ascii=False, indent=2)

    prompt = f"""Genera un email formal de seguimiento de fin de jornada del {fd_fmt}, proyecto LM5 Río Blanco, programa 3WLA.

Aquí están TODAS las actividades del día con su estado de ejecución:
{datos}

FORMATO EXACTO — SIGUE ESTE EJEMPLO SIN NINGUNA VARIACIÓN:

Estimados,

Junto con saludar, comparto el seguimiento de las actividades planificadas en el 3WLA para la jornada de hoy, con sus respectivas observaciones:

IIFF

Manejo de sitio (housekeeping) (Día) — Ejecutado: 61,66 HH
Instalaciones y Controles de Acceso Vial (Día) — Ejecutado: 15,86 HH

TRAMO 2

Traslado Tubería HDPE BODEGA - PTO TRABAJO (Día + Pendiente | 30,57 HH comprometidas): Causa del no cumplimiento. Responsable: Fe Grande.
Desfile de tuberías (4 un) (Día | 24,63 HH comprometidas): Causa del no cumplimiento. Responsable: Fe Grande.
Unión de tuberías tubo - Tubo (4 un) (Día + Pendiente | 166,22 HH comprometidas): Causa del no cumplimiento. Responsable: Fe Grande.

TRAMO 4

Excavación suelo Común (Día + Pendiente) — Ejecutado: 29,82 HH
Uniones tuberías HDPE: tubo-codo 90° (Solo Pendiente) — Ejecutado: 249,70 HH
Uniones tuberías HDPE: tubo-codo 45° (Día | 237,00 HH comprometidas): Causa del no cumplimiento. Responsable: Fe Grande.

AVANCE ACUMULADO S1 – 3WLA

Total ejecutado a la fecha: 1.259,26 HH de 4.946,88 HH comprometidas.
Total esperado a la fecha: 1.941,67 HH de 4.946,88 HH comprometidas.
Nota: El total de 4.946,88 HH se considera ajustado, dado que no incorpora las actividades de venteo ni los trabajos eléctricos del Tramo 1.

Saludos.

Atte.

JCP

---

REGLAS ABSOLUTAS — NO LAS IGNORES BAJO NINGÚN CONCEPTO:

1. CERO emojis (sin ✅ ni ❌ ni ningún otro)
2. CERO asteriscos, negritas, cursivas ni markdown de ningún tipo
3. CERO separadores --- entre secciones
4. El nombre del área va solo en su línea, sin ningún símbolo
5. Cada actividad va separada por una línea en blanco de la siguiente
6. Cada actividad sigue este formato exacto:
   - Si ejecutado=true:  Nombre actividad (Categoría) — Ejecutado: X,XX HH
   - Si ejecutado=false: Nombre actividad (Categoría | X,XX HH no ejecutadas): causa. Responsable: X.
     donde X,XX = campo hh_deficit (déficit neto acumulado real de la actividad)
7. Categoría según campo "categoria":
   - "Planificada del día"      → escribir "Día"
   - "Planificada + pendiente"  → escribir "Día + Pendiente"
   - "Solo pendiente"           → escribir "Solo Pendiente"
8. Números formato chileno: punto para miles, coma para decimal (1.259,26)
9. Incluye TODAS las actividades del JSON sin omitir ninguna
10. Solo incluir el área si tiene al menos una actividad
11. En AVANCE ACUMULADO S1 usa los valores reales: {hh_ej_t:,.2f} HH ejecutadas y {hh_esp_t:,.2f} HH esperadas de {hh_meta:,.2f} HH
12. Texto plano, limpio y formal. Sin ningún tipo de formato especial"""

    client = anthropic.Anthropic()
    resp = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=8000,
        temperature=0,
        messages=[{"role": "user", "content": prompt}]
    )
    return resp.content[0].text

# ─────────────────────────────────────────────
# INTERFAZ PRINCIPAL
# ─────────────────────────────────────────────
def main():
    # ── LOGIN ────────────────────────────────────
    APP_PASSWORD = os.environ.get("APP_PASSWORD", "3wla2026")

    if "autenticado" not in st.session_state:
        st.session_state["autenticado"] = False

    if not st.session_state["autenticado"]:
        st.markdown("## 🏗️ Sistema de Seguimiento Diario — 3WLA")
        st.caption("Proyecto LM5 Río Blanco | Creado por Joaquín Cuello Poblete")
        st.markdown("---")
        col_l1, col_l2, col_l3 = st.columns([1,2,1])
        with col_l2:
            st.markdown("### 🔒 Acceso restringido")
            pwd = st.text_input("Contraseña", type="password", placeholder="Ingresa la contraseña")
            if st.button("Ingresar", type="primary", use_container_width=True):
                if pwd == APP_PASSWORD:
                    st.session_state["autenticado"] = True
                    st.rerun()
                else:
                    st.error("Contraseña incorrecta")
        return

    # Header
    col_h1, col_h2 = st.columns([3, 1])
    with col_h1:
        st.markdown("## 🏗️ Sistema de Seguimiento Diario — 3WLA")
        st.caption("Proyecto LM5 Río Blanco | Creado por Joaquín Cuello Poblete")

    estado = cargar_estado()

    # ── CARGA DE TRISEMANAL ───────────────────
    with st.expander("📂 Cargar nuevo trisemanal (.xlsx)", expanded=not bool(estado)):
        uploaded = st.file_uploader("Selecciona el archivo", type=["xlsx"], label_visibility="collapsed")
        if uploaded:
            xlsx_bytes = uploaded.read()
            with open("trisemanal_temp.xlsx", "wb") as f:
                f.write(xlsx_bytes)
            with st.spinner("Leyendo trisemanal..."):
                try:
                    tri = leer_trisemanal("trisemanal_temp.xlsx")
                    # Archivar trisemanal anterior en histórico si existe
                    historico = estado.get("historico", []) if estado else []
                    if estado and estado.get("trisemanal") and estado.get("registro"):
                        tri_ant = estado["trisemanal"]
                        acu_ant = calcular_acumulado(estado, tri_ant["fechas_s1"][-1])
                        historico.append({
                            "num_trisemanal": tri_ant.get("num_trisemanal","—"),
                            "periodo": f"{tri_ant['fechas_s1'][0]} → {tri_ant['fechas_s1'][-1]}",
                            "fechas_s1": tri_ant["fechas_s1"],
                            "hh_meta": tri_ant["hh_totales_s1"],
                            "hh_ejecutadas": acu_ant["hh_ejecutadas"],
                            "hh_esperadas": acu_ant["hh_esperadas"],
                            "por_area": acu_ant["por_area"],
                            "por_resp": acu_ant["por_resp"],
                            "registro": estado["registro"],
                            "actividades": tri_ant["actividades"],
                        })
                    estado_nuevo = {"trisemanal": tri, "registro": {}, "historico": historico}
                    guardar_estado(estado_nuevo)
                    estado = estado_nuevo
                    st.success(f"✅ Trisemanal N°{tri['num_trisemanal']} cargado — "
                               f"S1: {tri['fechas_s1'][0]} → {tri['fechas_s1'][-1]} — "
                               f"Meta: {tri['hh_totales_s1']:,.1f} HH")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error leyendo trisemanal: {e}. Detalle: hh_totales={tri.get('hh_totales_s1','?') if 'tri' in dir() else '?'}")

    if not estado:
        st.info("👆 Carga el trisemanal para comenzar.")
        return

    tri = estado["trisemanal"]

    # Botón para archivar período actual manualmente
    with st.expander("📦 Archivar período actual y empezar nuevo", expanded=False):
        st.caption("Guarda el período actual en el historial y limpia el registro para recibir el nuevo trisemanal.")
        if st.button("📦 Archivar período y limpiar registro", type="secondary", use_container_width=True):
            historico = estado.get("historico", [])
            if estado.get("registro"):
                tri_act = estado["trisemanal"]
                acu_act = calcular_acumulado(estado, tri_act["fechas_s1"][-1])
                historico.append({
                    "num_trisemanal": tri_act.get("num_trisemanal","—"),
                    "periodo": f"{tri_act['fechas_s1'][0]} → {tri_act['fechas_s1'][-1]}",
                    "fechas_s1": tri_act["fechas_s1"],
                    "hh_meta": tri_act["hh_totales_s1"],
                    "hh_ejecutadas": acu_act["hh_ejecutadas"],
                    "hh_esperadas": acu_act["hh_esperadas"],
                    "por_area": acu_act["por_area"],
                    "por_resp": acu_act["por_resp"],
                    "registro": estado["registro"],
                    "actividades": tri_act["actividades"],
                })
                estado["historico"] = historico
                estado["registro"]  = {}
                guardar_estado(estado)
                st.success("✅ Período archivado correctamente. Ahora carga el nuevo trisemanal.")
                st.rerun()
            else:
                st.warning("No hay registro que archivar.")

    # Info rápida
    with col_h2:
        acu_rapido = calcular_acumulado(estado, date.today().isoformat())
        pct = round(acu_rapido["hh_ejecutadas"] / tri["hh_totales_s1"] * 100, 1) if tri["hh_totales_s1"] else 0
        st.metric("Avance S1", f"{pct}%",
                  f"{acu_rapido['hh_ejecutadas']:,.0f} / {tri['hh_totales_s1']:,.0f} HH")

    # Selector de fecha
    st.markdown("---")
    col_f1, col_f2 = st.columns([2, 3])
    with col_f1:
        fecha_min = date.fromisoformat(tri["fechas_s1"][0])
    fecha_max = date.fromisoformat(tri["fechas_s1"][-1])
    fecha_default = min(date.today(), fecha_max)
    fecha_default = max(fecha_default, fecha_min)
    fecha_sel = st.date_input(
            "📅 Fecha de trabajo",
            value=fecha_default,
            min_value=fecha_min,
            max_value=fecha_max,
        )
    fecha_str  = fecha_sel.isoformat()
    fecha_fmt  = fecha_sel.strftime("%A %d de %B").capitalize()
    with col_f2:
        st.markdown(f"<br><b style='font-size:18px'>{fecha_fmt}</b> — "
                    f"Trisemanal N°{tri['num_trisemanal']} | "
                    f"S1: {tri['fechas_s1'][0]} → {tri['fechas_s1'][-1]}",
                    unsafe_allow_html=True)

    acts_hoy   = actividades_del_dia(estado, fecha_str)
    pendientes = pendientes_acumulados(estado, fecha_str)

    if not acts_hoy:
        st.warning("No hay actividades programadas para esta fecha en S1.")
        panel_acumulado(estado, fecha_str, tab_key="noacts")
        return

    # ── TABS PRINCIPALES ─────────────────────
    tab_actual, tab_historico = st.tabs(["📋  TRISEMANAL ACTUAL", "📈  COMPILADO HISTÓRICO"])

    with tab_historico:
        historico = estado.get("historico", [])
        if not historico:
            st.info("Aún no hay períodos anteriores archivados. El historial se construye automáticamente al cargar un nuevo trisemanal.")
        else:
            import pandas as pd

            # ── RESUMEN GENERAL ──────────────────────────────
            st.markdown("### Resumen acumulado de todos los períodos")
            total_ej  = sum(h["hh_ejecutadas"] for h in historico)
            total_meta = sum(h["hh_meta"] for h in historico)
            total_esp = sum(h["hh_esperadas"] for h in historico)
            pct_total = round(total_ej / total_meta * 100, 1) if total_meta else 0

            c1, c2, c3, c4 = st.columns(4)
            with c1: st.metric("✅ HH Ejecutadas", f"{total_ej:,.1f}")
            with c2: st.metric("📅 HH Esperadas", f"{total_esp:,.1f}")
            with c3: st.metric("📊 HH Meta total", f"{total_meta:,.1f}")
            with c4: st.metric("🎯 Avance global", f"{pct_total}%")
            st.progress(pct_total / 100, text=f"Avance acumulado: {pct_total}%")

            # ── TABLA POR PERÍODO ────────────────────────────
            st.markdown("---")
            st.markdown("### Detalle por período")
            rows_hist = []
            for h in historico:
                pct_h = round(h["hh_ejecutadas"]/h["hh_meta"]*100,1) if h["hh_meta"] else 0
                rows_hist.append({
                    "Período":         h["periodo"],
                    "Trisemanal":      f"N°{h['num_trisemanal']}",
                    "Meta (HH)":       round(h["hh_meta"],1),
                    "Esperadas (HH)":  round(h["hh_esperadas"],1),
                    "Ejecutadas (HH)": round(h["hh_ejecutadas"],1),
                    "Déficit (HH)":    round(h["hh_esperadas"]-h["hh_ejecutadas"],1),
                    "% Cumpl.":        f"{pct_h}%",
                })
            df_hist = pd.DataFrame(rows_hist)
            st.dataframe(df_hist, use_container_width=True, hide_index=True)

            # ── GRÁFICO COMPARATIVO ──────────────────────────
            try:
                import plotly.graph_objects as go
                periodos = [h["periodo"].split(" → ")[0] for h in historico]
                vals_meta = [h["hh_meta"] for h in historico]
                vals_ej   = [h["hh_ejecutadas"] for h in historico]
                vals_esp  = [h["hh_esperadas"] for h in historico]

                fig_hist = go.Figure()
                fig_hist.add_trace(go.Bar(name="Meta", x=periodos, y=vals_meta,
                    marker_color="#E2E8F0", text=[f"{v:,.0f}" for v in vals_meta], textposition="outside"))
                fig_hist.add_trace(go.Bar(name="Esperado", x=periodos, y=vals_esp,
                    marker_color="#93C5FD", text=[f"{v:,.0f}" for v in vals_esp], textposition="outside"))
                fig_hist.add_trace(go.Bar(name="Ejecutado", x=periodos, y=vals_ej,
                    marker_color="#2563EB", text=[f"{v:,.0f}" for v in vals_ej], textposition="outside"))
                fig_hist.update_layout(title="HH por período — Meta vs Esperado vs Ejecutado",
                    barmode="group", height=420, legend=dict(orientation="h", y=-0.2),
                    margin=dict(t=50,b=60))
                st.plotly_chart(fig_hist, use_container_width=True, key="chart_hist_1")
            except Exception:
                pass

            # ── DETALLE POR ÁREA ACUMULADO ───────────────────
            st.markdown("---")
            st.markdown("### Detalle por área — acumulado histórico")
            areas_acum = {}
            for h in historico:
                for area, v in h.get("por_area", {}).items():
                    if area not in areas_acum:
                        areas_acum[area] = {"esp":0,"ej":0}
                    areas_acum[area]["esp"] += v.get("esp",0)
                    areas_acum[area]["ej"]  += v.get("ej",0)

            if areas_acum:
                rows_area_h = []
                for area, v in sorted(areas_acum.items()):
                    pct_a = round(v["ej"]/v["esp"]*100,1) if v["esp"]>0 else 0
                    rows_area_h.append({"Área": area, "HH Esperadas": round(v["esp"],1),
                        "HH Ejecutadas": round(v["ej"],1), "% Cumpl.": f"{pct_a}%"})
                st.dataframe(pd.DataFrame(rows_area_h), use_container_width=True, hide_index=True)

            # ── RESPONSABILIDADES ACUMULADAS ─────────────────
            st.markdown("---")
            st.markdown("### HH no ejecutadas por responsable — acumulado histórico")
            resp_acum = {}
            for h in historico:
                for resp, v in h.get("por_resp", {}).items():
                    resp_acum[resp] = resp_acum.get(resp,0) + v
            if resp_acum:
                total_r = sum(resp_acum.values())
                rows_resp_h = [{"Responsable":r, "HH no ejecutadas":round(v,1),
                    "% del total":f"{round(v/total_r*100,1)}%"} for r,v in sorted(resp_acum.items(),key=lambda x:-x[1])]
                st.dataframe(pd.DataFrame(rows_resp_h), use_container_width=True, hide_index=True)

            # ── ACTIVIDADES NO CUMPLIDAS HISTÓRICAS ──────────
            st.markdown("---")
            st.markdown("### Actividades no cumplidas — todos los períodos")
            rows_hist_acts = []
            for h in historico:
                fechas_h = [date.fromisoformat(d) for d in h["fechas_s1"]]
                for a in h.get("actividades",[]):
                    if not a["inicio"] or not a["termino"]: continue
                    ini_h = date.fromisoformat(a["inicio"])
                    ter_h = date.fromisoformat(a["termino"])
                    hh_esp_h = sum(a["hh_dia"] for fd in fechas_h if ini_h<=fd<=ter_h)
                    hh_ej_h  = sum((h["registro"].get(fd.isoformat(),{}).get(str(a["corr"]),{}).get("cant_ejecutada",0) or 0)*a["rendimiento"]
                                   for fd in fechas_h)
                    deficit_h = round(hh_esp_h - hh_ej_h, 1)
                    if deficit_h > 0.5:
                        resp_h = ""
                        for fd in sorted(fechas_h, reverse=True):
                            reg_h = h["registro"].get(fd.isoformat(),{}).get(str(a["corr"]))
                            if reg_h and reg_h.get("responsable"):
                                resp_h = reg_h["responsable"]; break
                        rows_hist_acts.append({"Período": h["periodo"].split(" → ")[0],
                            "Área": a["area"], "Actividad": a["nombre"],
                            "HH Déficit": deficit_h, "Responsable": resp_h})
            if rows_hist_acts:
                df_ha = pd.DataFrame(sorted(rows_hist_acts, key=lambda x:-x["HH Déficit"]))
                st.dataframe(df_ha, use_container_width=True, hide_index=True,
                    column_config={"HH Déficit": st.column_config.NumberColumn(format="%.1f HH")})

    with tab_actual:
        tab1, tab2 = st.tabs(["🌅  INICIO DEL DÍA", "🌆  FIN DEL DÍA"])

    # ══════════════════════════════════════════
    # TAB 1 — INICIO DEL DÍA
    # ══════════════════════════════════════════
    with tab1:
        st.markdown("### ¿Qué debe ejecutarse hoy?")
        st.caption("Incluye lo programado para hoy más lo pendiente de días anteriores.")

        # Actividades agrupadas por área
        areas = sorted(set(a["area"] for a in acts_hoy))
        total_hh_hoy = 0

        for area in areas:
            st.markdown(f"**{area}**")
            acts_area = [a for a in acts_hoy if a["area"] == area]
            for a in acts_area:
                pend       = pendientes.get(a["corr"], 0)
                pend_positivo = max(pend, 0.0)
                # Si la actividad está en rango hoy: cant_dia + pendiente
                # Si está fuera de rango (solo aparece por deuda): solo pendiente
                fd_act = date.fromisoformat(fecha_str)
                ini_act = date.fromisoformat(a["inicio"]) if a["inicio"] else fd_act
                ter_act = date.fromisoformat(a["termino"]) if a["termino"] else fd_act
                en_rango_hoy = ini_act <= fd_act <= ter_act
                if en_rango_hoy:
                    cant_total = round(a["cant_dia"] + pend_positivo, 4)
                else:
                    cant_total = round(pend_positivo, 4)
                hh_total   = round(cant_total * a["rendimiento"], 2)
                total_hh_hoy += hh_total

                pend_tag = (f'<span class="pendiente-tag">+ {round(pend,4)} {a["unidad"]} pendiente</span>'
                            if pend > 0 else "")
                st.markdown(
                    f'<div class="act-card">'
                    f'<div class="act-nombre">{a["nombre"]}{pend_tag}</div>'
                    f'<div class="act-meta">'
                    f'<b>{cant_total} {a["unidad"]}</b> &nbsp;→&nbsp; '
                    f'<b style="color:#1d4ed8">{hh_total} HH</b>'
                    f'</div></div>',
                    unsafe_allow_html=True
                )

        st.markdown(f"**Total comprometido hoy: {total_hh_hoy:,.2f} HH**")
        st.markdown("---")

        if st.button("✉️ Generar email de inicio del día", type="primary", use_container_width=True):
            with st.spinner("Generando email..."):
                try:
                    email = generar_email_inicio(estado, fecha_str, acts_hoy, pendientes)
                    st.session_state["email_inicio"] = email
                except Exception as e:
                    st.error(f"Error: {e}")

        if "email_inicio" in st.session_state:
            st.markdown("#### 📧 Copia este email:")
            st.text_area("", value=st.session_state["email_inicio"],
                         height=380, key="ta_inicio", label_visibility="collapsed")

        panel_acumulado(estado, fecha_str, tab_key="inicio")

    # ══════════════════════════════════════════
    # TAB 2 — FIN DEL DÍA
    # ══════════════════════════════════════════
    with tab2:
        st.markdown("### ¿Qué se ejecutó hoy?")
        st.caption("Ingresa la cantidad real ejecutada por actividad.")

        registro_dia    = estado.get("registro", {}).get(fecha_str, {})
        registro_nuevo  = {}
        ya_registrado   = bool(registro_dia)

        if ya_registrado:
            st.success(f"✅ Ya existe un registro para {fecha_fmt}. Puedes editarlo.")

        areas = sorted(set(a["area"] for a in acts_hoy))

        # Leyenda de categorías
        st.markdown(
            '<div style="display:flex;gap:16px;margin-bottom:12px;flex-wrap:wrap">'
            '<span style="font-size:12px">🔵 <b>Planificada del día</b> — solo lo de hoy</span>'
            '<span style="font-size:12px">🟡 <b>Planificada + pendiente</b> — hoy más deuda anterior</span>'
            '<span style="font-size:12px">🔴 <b>Solo pendiente</b> — deuda de días anteriores</span>'
            '</div>',
            unsafe_allow_html=True
        )

        for area in areas:
            st.markdown(f"**{area}**")
            acts_area = [a for a in acts_hoy if a["area"] == area]

            for a in acts_area:
                pend         = pendientes.get(a["corr"], 0)
                pend_positivo = max(pend, 0.0)
                fd_act       = date.fromisoformat(fecha_str)
                ini_act      = date.fromisoformat(a["inicio"]) if a["inicio"] else fd_act
                ter_act      = date.fromisoformat(a["termino"]) if a["termino"] else fd_act
                en_rango_hoy = ini_act <= fd_act <= ter_act

                # Determinar categoría
                if en_rango_hoy and pend_positivo > 0.0001:
                    categoria    = "🟡"
                    cat_label    = "Planificada + pendiente"
                    cat_color    = "#92400e"
                    cat_bg       = "#fef3c7"
                    cant_dia_hoy = round(a["cant_dia"], 4)
                    cant_total   = round(a["cant_dia"] + pend_positivo, 4)
                    detalle      = (f"{cant_dia_hoy} {a['unidad']} hoy"
                                   f" + {round(pend_positivo,4)} {a['unidad']} pendiente"
                                   f" = {cant_total} {a['unidad']}")
                elif en_rango_hoy:
                    categoria    = "🔵"
                    cat_label    = "Planificada del día"
                    cat_color    = "#1e3a8a"
                    cat_bg       = "#dbeafe"
                    cant_total   = round(a["cant_dia"], 4)
                    detalle      = f"{cant_total} {a['unidad']} hoy"
                else:
                    categoria    = "🔴"
                    cat_label    = "Solo pendiente"
                    cat_color    = "#991b1b"
                    cat_bg       = "#fee2e2"
                    cant_total   = round(pend_positivo, 4)
                    detalle      = f"{cant_total} {a['unidad']} pendiente de días anteriores"

                hh_total  = round(cant_total * a["rendimiento"], 2)
                corr_str  = str(a["corr"])
                prev_ej   = max(registro_dia.get(corr_str, {}).get("cant_ejecutada", cant_total), 0.0)

                with st.container():
                    st.markdown(
                        f'<div class="act-card" style="border-left:4px solid {cat_color}">'
                        f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:4px">'
                        f'<span style="font-size:13px">{categoria}</span>'
                        f'<span style="background:{cat_bg};color:{cat_color};font-size:11px;'
                        f'font-weight:600;padding:2px 8px;border-radius:20px">{cat_label}</span>'
                        f'</div>'
                        f'<div class="act-nombre">{a["nombre"]}</div>'
                        f'<div class="act-meta">{detalle} → '
                        f'<b style="color:#1d4ed8">{hh_total} HH</b></div>'
                        f'</div>',
                        unsafe_allow_html=True
                    )

                    c1, c2 = st.columns([3, 2])
                    with c1:
                        step = max(round(a["cant_dia"] / 4, 4), 0.0001)
                        max_val = float(max(cant_total, a["cant_dia"]) * 3)
                        safe_prev = min(float(prev_ej), max_val)
                        ejecutado = st.number_input(
                            f"¿Cuánto {a['unidad']} se ejecutó?",
                            min_value=0.0,
                            max_value=max_val,
                            value=safe_prev,
                            step=float(step),
                            key=f"ej_{a['corr']}_{fecha_str}",
                            format="%.4f"
                        )
                    with c2:
                        hh_ej_calc = round(ejecutado * a["rendimiento"], 2)
                        delta_col  = "normal" if ejecutado >= cant_total else "inverse"
                        st.metric("HH ejecutadas", f"{hh_ej_calc:,.2f}",
                                  f"{round(hh_ej_calc - hh_total, 2):+.2f} HH",
                                  delta_color=delta_col)

                    causa       = ""
                    responsable = ""
                    deficit_neto = round(cant_total - ejecutado, 4)
                    if deficit_neto > 0.0001:
                        col_c1, col_c2 = st.columns(2)
                        with col_c1:
                            causa = st.text_input(
                                "Causa del no cumplimiento",
                                value=registro_dia.get(corr_str, {}).get("causa", ""),
                                key=f"causa_{a['corr']}_{fecha_str}",
                                placeholder="Ej: Falta certificación GPRO"
                            )
                        with col_c2:
                            opciones_resp = ["Fe Grande", "R&Q Ingeniería", "CODELCO", "Externo", "Otro"]
                            prev_resp = registro_dia.get(corr_str, {}).get("responsable", "Fe Grande")
                            idx_resp  = opciones_resp.index(prev_resp) if prev_resp in opciones_resp else 0
                            responsable = st.selectbox(
                                "Responsable",
                                options=opciones_resp,
                                index=idx_resp,
                                key=f"resp_{a['corr']}_{fecha_str}"
                            )

                    registro_nuevo[corr_str] = {
                        "corr":           a["corr"],
                        "nombre":         a["nombre"],
                        "area":           a["area"],
                        "unidad":         a["unidad"],
                        "rendimiento":    a["rendimiento"],
                        "cant_dia_base":  a["cant_dia"],
                        "cant_esperada":  cant_total,
                        "cant_ejecutada": ejecutado,
                        "hh_esperadas":   hh_total,
                        "hh_ejecutadas":  hh_ej_calc,
                        "categoria":      cat_label,
                        "causa":          causa,
                        "responsable":    responsable,
                    }

        st.markdown("---")
        col_b1, col_b2 = st.columns(2)
        with col_b1:
            if st.button("💾 Guardar registro", use_container_width=True):
                if "registro" not in estado:
                    estado["registro"] = {}
                estado["registro"][fecha_str] = registro_nuevo
                guardar_estado(estado)
                st.success("✅ Registro guardado")
                st.rerun()

        with col_b2:
            if st.button("✉️ Guardar y generar email de cierre", type="primary", use_container_width=True):
                if "registro" not in estado:
                    estado["registro"] = {}
                estado["registro"][fecha_str] = registro_nuevo
                guardar_estado(estado)
                with st.spinner("Generando email..."):
                    try:
                        email = generar_email_cierre(estado, fecha_str, registro_nuevo)
                        st.session_state["email_cierre"] = email
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error: {e}")

        if "email_cierre" in st.session_state:
            st.markdown("#### 📧 Copia este email:")
            st.text_area("", value=st.session_state["email_cierre"],
                         height=420, key="ta_cierre", label_visibility="collapsed")

        panel_acumulado(estado, fecha_str, tab_key="fin")

if __name__ == "__main__":
    main()
